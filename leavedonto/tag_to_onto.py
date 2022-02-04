from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Protection
from openpyxl.utils import coordinate_to_tuple

from .trie import OntTrie
from .dataval import DataVal
from .utils import resize_sheet


def tagged_to_trie(tagged, onto_basis):
    trie = OntTrie()
    trie.legend = onto_basis.ont.legend
    for word, pos, level in tagged:
        found = onto_basis.ont.find_entries(prefix=pos, lemma=word)
        if found:
            for path, entries in found:
                for e in entries:
                    found_level = onto_basis.get_field_value(e, "level")
                    if found_level == level:
                        trie.add(path, e)
        else:
            path = [pos, "to_organize"]
            parts = {"word": word, "POS": pos, "level": level}
            entry = [parts[l] if l in parts else "" for l in onto_basis.ont.legend]
            trie.add(path, entry)
    return trie


def get_entries(in_file):
    wb = load_workbook(in_file)
    ws = wb.active

    # from sheet to list of lists
    tagged = []
    max_row, max_col = coordinate_to_tuple(ws.dimensions.split(":")[1])
    for r in range(1, max_row + 1, 4):
        for col in range(1, max_col + 1):
            # ignoring the first column containing the numbers
            word = ws.cell(r, col).value
            pos = ws.cell(r + 1, col).value
            level = ws.cell(r + 2, col).value
            entry = (word, pos, level)
            if word and pos and level and entry not in tagged:
                tagged.append(entry)

    return tagged


def generate_to_tag(in_file, onto, pos_list, levels, l_colors, out_file=None, fields=dict):
    # first load all the ontos you need in OntoManager, then run
    font = "Jomolhari"
    ft_words = Font(font, size=17, color="000c1d91")
    ft_pos = Font(font, size=13, color="004e4f54")
    ft_level = Font(size=11, color="004e4f54")
    new_bgcolor = PatternFill("solid", fgColor="0090f0a9")
    alignmnt = Alignment(horizontal="left", vertical="center")

    wb = Workbook()
    wb.remove(wb.get_sheet_by_name("Sheet"))

    # read input file into rows
    lines = in_file.read_text().lstrip("\ufeff").split("\n")
    rows = rows_from_lines(lines)

    # prepare data validation for POS and levels
    dv = DataVal(wb)
    dv.add_validator("POS", pos_list)
    dv.add_validator("level", levels)

    # create sheet and fill it
    sheet_name = in_file.stem.split("_")[0]
    ws = wb.create_sheet(title=sheet_name)
    ws.protection.sheet = True
    for n, r in enumerate(rows):
        row = n * 4 + 1
        pos_row = row + 1
        level_row = pos_row + 1
        for m, el in enumerate(r):
            col = m + 1

            # check if word exists in onto
            found = onto.find_word(el)
            found_pos = found[0][0][0] if found else None
            entries = found[0][1] if found else None

            # add word to spreadsheet
            word_cell = ws.cell(row=row, column=col)
            word_cell.value = el
            word_cell.font = ft_words
            word_cell.alignment = alignmnt

            # add POS
            pos_cell = ws.cell(row=pos_row, column=col)
            pos_cell.protection = Protection(locked=False)
            pos_cell.value = found_pos if found_pos else ""
            pos_cell.font = ft_pos
            pos_cell.alignment = alignmnt
            dv.add_val_to_cell(
                val_name="POS", sheet_name=sheet_name, row=pos_row, col=col
            )
            if not entries:
                pos_cell.fill = new_bgcolor

            # add level
            level_cell = ws.cell(row=level_row, column=col)
            level_cell.protection = Protection(locked=False)
            level_cell.value = (
                onto.get_field_value(entries[0], "level") if entries else fields["level"]
            )
            level_cell.font = ft_level
            level_cell.alignment = alignmnt
            dv.add_val_to_cell(
                val_name="level", sheet_name=sheet_name, row=level_row, col=col
            )
            if not entries:
                level_cell.fill = new_bgcolor

        # set row height for each group of "word, POS and level"
        ws.row_dimensions[row].height = 30
        ws.row_dimensions[pos_row].height = 15
        ws.row_dimensions[level_row].height = 15
        ws.row_dimensions[
            level_row + 1
        ].height = 30  # size of empty row between two lines

    resize_sheet(ws, mode="width")

    if not out_file:
        out_file = in_file.parent / (in_file.stem + "_totag.xlsx")

    wb.save(out_file)


def generate_to_tag_chunks(chunks, config, onto, pos_list, levels, l_colors, out_file, fields=dict):
    # first load all the ontos you need in OntoManager, then run
    font = "Jomolhari"
    ft_words = Font(font, size=13, color="004e4f54")
    ft_pos = Font(font, size=17, color="000c1d91")
    ft_level = Font(size=11, color="004e4f54")
    new_bgcolor = PatternFill("solid", fgColor="0090f0a9")
    alignmnt = Alignment(horizontal="left", vertical="center")

    if out_file.is_file():
        wb = load_workbook(out_file)
    else:
        wb = Workbook()
        wb.remove(wb.get_sheet_by_name("Sheet"))

    # prepare data validation for POS and levels
    dv = DataVal(wb)
    dv.add_validator("POS", pos_list)
    dv.add_validator("level", levels)

    # create sheet
    sheet_name = out_file.stem.split("_")[0]
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(title=sheet_name)
    else:
        ws = wb[sheet_name]
    ws.protection.sheet = True

    # process chunks
    has_processed = False  # ensures only one chunk is processed at a time
    for c_count, chunk in chunks.items():
        if config[c_count] == 'todo' and not has_processed:
            has_processed = True
            config[c_count] = 'done'

            # read input file into rows
            rows = rows_from_lines(chunk)
            row_start = c_count * 4 * 4
            for n, r in enumerate(rows):
                row = row_start + n * 4 + 1
                pos_row = row + 1
                level_row = pos_row + 1
                for m, el in enumerate(r):
                    col = m + 1

                    # check if word exists in onto
                    found = onto.find_word(el)
                    found_pos = found[0][0][0] if found else None
                    entries = found[0][1] if found else None

                    # add word to spreadsheet
                    word_cell = ws.cell(row=row, column=col)
                    word_cell.value = el
                    word_cell.font = ft_words
                    word_cell.alignment = alignmnt

                    # add POS
                    pos_cell = ws.cell(row=pos_row, column=col)
                    pos_cell.protection = Protection(locked=False)
                    pos_cell.value = found_pos if found_pos else ""
                    pos_cell.font = ft_pos
                    pos_cell.alignment = alignmnt
                    dv.add_val_to_cell(
                        val_name="POS", sheet_name=sheet_name, row=pos_row, col=col
                    )
                    if not entries:
                        pos_cell.fill = new_bgcolor

                    # add level
                    level_cell = ws.cell(row=level_row, column=col)
                    level_cell.protection = Protection(locked=False)
                    level_cell.value = (
                        onto.get_field_value(entries[0], "level") if entries else fields["level"]
                    )
                    level_cell.font = ft_level
                    level_cell.alignment = alignmnt
                    dv.add_val_to_cell(
                        val_name="level", sheet_name=sheet_name, row=level_row, col=col
                    )
                    if not entries:
                        level_cell.fill = new_bgcolor

                # set row height for each group of "word, POS and level"
                ws.row_dimensions[row].height = 20
                ws.row_dimensions[pos_row].height = 30
                ws.row_dimensions[level_row].height = 15
                ws.row_dimensions[
                    level_row + 1
                ].height = 30  # size of empty row between two lines

    resize_sheet(ws, mode="width")

    wb.save(out_file)
    return config


def rows_from_lines(lines, mode='lines'):
    row_size = 12

    if mode == 'lines':
        words = []
        for l in lines:
            words.extend(l.split(" "))
    else:
        words = lines

    rows = []
    cur_row = []
    while words:
        cur_row.append(words.pop(0))

        if len(cur_row) == row_size:
            rows.append(cur_row)
            cur_row = []
    if cur_row:
        rows.append(cur_row)

    return rows
